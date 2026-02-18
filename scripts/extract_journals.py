"""
从Excel文件中提取标黄的期刊名称，生成 JSON 配置文件。
用法：python src/extract_journals.py --input <excel文件路径> [--output journals.json]

支持 .xlsx (openpyxl) 和 .xls (xlrd) 两种格式。
"""
import argparse
import json
import sys
import os


def extract_from_xlsx(excel_path: str) -> list:
    """从 .xlsx 文件中提取标黄期刊"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(excel_path, data_only=True)
    highlighted_journals = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📄 正在扫描工作表: {sheet_name}")
        print(f"   总行数: {ws.max_row}, 总列数: {ws.max_column}")

        # 查找期刊名列
        journal_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and any(kw in str(header).lower() for kw in ["journal", "title", "期刊"]):
                journal_col = col
                print(f"   ✅ 找到期刊名列: 第{col}列 (标题: {header})")
                break
        if journal_col is None:
            journal_col = 2
            print(f"   ⚠️ 未找到明确的期刊列标题，默认使用第{journal_col}列")

        yellow_count = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=journal_col)
            is_highlighted = False
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                fill = c.fill
                if fill and fill.fgColor and fill.fgColor.rgb:
                    color = str(fill.fgColor.rgb)
                    if color in ("FFFFFF00", "00FFFF00", "FFFFFFCC", "FFFFF200"):
                        is_highlighted = True
                        break
                    if len(color) >= 6:
                        try:
                            hex_color = color[-6:]
                            r = int(hex_color[0:2], 16)
                            g = int(hex_color[2:4], 16)
                            b = int(hex_color[4:6], 16)
                            if r > 200 and g > 180 and b < 100:
                                is_highlighted = True
                                break
                        except ValueError:
                            pass

            if is_highlighted and cell.value:
                journal_name = str(cell.value).strip()
                if journal_name and journal_name.lower() not in ["journal", "full journal title", "title", "期刊名"]:
                    highlighted_journals.append(journal_name)
                    yellow_count += 1

        print(f"   🔍 找到 {yellow_count} 个标黄的期刊")
    return highlighted_journals


def extract_from_xls(excel_path: str) -> list:
    """从 .xls 文件中提取标黄期刊"""
    try:
        import xlrd
    except ImportError:
        print("❌ 需要安装 xlrd: pip install xlrd")
        sys.exit(1)

    wb = xlrd.open_workbook(excel_path, formatting_info=True)
    highlighted_journals = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        print(f"\n📄 正在扫描工作表: {ws.name}")
        print(f"   总行数: {ws.nrows}, 总列数: {ws.ncols}")

        # 查找期刊名列 — 扫描前10行寻找列标题
        # 注意：要跳过 "Journal Data Filtered By..." 这种长描述文本
        journal_col = None
        header_row = 0
        for row in range(min(10, ws.nrows)):
            for col in range(ws.ncols):
                header = str(ws.cell_value(row, col)).strip()
                if not header or len(header) > 50:
                    continue  # 跳过空值和长描述文本
                header_lower = header.lower()
                # 精确匹配已知的期刊列标题
                if header_lower in ["full journal title", "journal title", "journal name", "期刊名", "期刊"]:
                    journal_col = col
                    header_row = row
                    print(f"   ✅ 找到期刊名列: 第{col+1}列, 第{row+1}行 (标题: {header})")
                    break
            if journal_col is not None:
                break
                
        if journal_col is None:
            journal_col = 1  # 默认第2列 (0-indexed)
            header_row = 1   # 假设第2行是标题行
            print(f"   ⚠️ 未找到明确的期刊列标题，默认使用第{journal_col+1}列")

        # 获取XF格式信息以检查背景色
        yellow_count = 0
        for row in range(header_row + 1, ws.nrows):  # 跳过标题行
            is_highlighted = False
            for col in range(ws.ncols):
                try:
                    xf_index = ws.cell_xf_index(row, col)
                    xf = wb.xf_list[xf_index]
                    bg_color_idx = xf.background.pattern_colour_index
                    # xlrd中，黄色的颜色索引通常是: 13 (Yellow), 43 (Light Yellow), 51 等
                    if bg_color_idx in (13, 43, 51, 44, 34, 52, 5, 6):
                        is_highlighted = True
                        break
                    # 尝试通过 colour_map 获取 RGB
                    color_map = wb.colour_map.get(bg_color_idx)
                    if color_map:
                        r, g, b = color_map
                        if r and g and b and r > 200 and g > 180 and b < 100:
                            is_highlighted = True
                            break
                except Exception:
                    pass

            if is_highlighted:
                cell_value = ws.cell_value(row, journal_col)
                if cell_value:
                    journal_name = str(cell_value).strip()
                    # 跳过标题行、纯数字（Rank列）、空值
                    if journal_name and journal_name.lower() not in ["journal", "full journal title", "title", "期刊名", "rank"]:
                        # 跳过纯数字（可能是Rank列的值）
                        try:
                            float(journal_name)
                            continue  # 是数字，跳过
                        except ValueError:
                            pass
                        highlighted_journals.append(journal_name)
                        yellow_count += 1

        print(f"   🔍 找到 {yellow_count} 个标黄的期刊")
    return highlighted_journals


def extract_highlighted_journals(excel_path: str, output_path: str = "journals.json"):
    ext = os.path.splitext(excel_path)[1].lower()

    if ext == ".xlsx":
        highlighted_journals = extract_from_xlsx(excel_path)
    elif ext == ".xls":
        highlighted_journals = extract_from_xls(excel_path)
    else:
        print(f"❌ 不支持的文件格式: {ext}，请使用 .xls 或 .xlsx 文件")
        sys.exit(1)

    # 去重
    unique_journals = list(dict.fromkeys(highlighted_journals))

    result = {
        "target_journals": unique_journals,
        "_description": "从Excel中提取的标黄期刊列表，用于 OptoAgent 的 tracking_sources.json"
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=4, ensure_ascii=False)

    print(f"\n✅ 共提取 {len(unique_journals)} 个期刊，已保存到: {output_path}")
    print("\n提取到的期刊列表:")
    for idx, j in enumerate(unique_journals, 1):
        print(f"  {idx}. {j}")

    return unique_journals


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="从Excel中提取标黄的期刊名")
    parser.add_argument("--input", "-i", required=True, help="Excel文件路径 (.xls 或 .xlsx)")
    parser.add_argument("--output", "-o", default="journals.json", help="输出JSON文件路径")
    args = parser.parse_args()

    extract_highlighted_journals(args.input, args.output)
